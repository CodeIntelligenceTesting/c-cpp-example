#!/usr/bin/env python3

import copy
import json
import openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, FormulaRule
import os
import re
import subprocess
import sys

def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def parse_function(function):
    eprint(function)
    # Remove whitespace and trailing semicolon
    function = function.strip().strip(';')
    # For parsing make sure there is no space before the parenthesis
    function = re.sub(r' *\(', '(', function)
    # Replace (void) notation with empty parentheses
    function = re.sub(r'\( *void *\)', '()', function)
    # Parse return type, function name and parameter string
    match = re.match(r'^(.*[\* ])(\w+)\((.*)\)$', function)
    if not match:
        raise Exception(f'Function did not match regex: {function}')
    func_data = {}
    func_data['name'] = match.group(2).strip()
    # Parse modifiers and return type
    mod_ret_str = match.group(1).strip()
    mod_ret = mod_ret_str.split(' ')
    mod_ret = [mr.strip() for mr in mod_ret if mr]
    # Fix pointer return type which could be split up
    if len(mod_ret) >= 2 and mod_ret[-1] == '*':
        mod_ret[-2] += ' *'
        del mod_ret[-1]
    func_data['return_type'] = mod_ret[-1]
    func_data['attributes'] = mod_ret[:-1]
    # Parse parameters
    func_data['params'] = {}
    if match.group(3).strip() != '':
        params_str = [param_str.strip() for param_str in match.group(3).split(',')]
        for i, param_str in enumerate(params_str):
            eprint(param_str)
            pmatch = re.fullmatch(r'(.*[\* ])([\w\[\]]+)', param_str)
            if pmatch:
                func_data['params'][pmatch.group(2).strip()] = pmatch.group(1).strip()
            else:
                ptmatch = re.fullmatch(r'([\w]+(?: *)?(?:\**)?)', param_str)
                if ptmatch:
                    func_data['params'][f'param{i}'] = ptmatch.group(1).strip()
                else:
                    raise Exception(f'Parameter in function "{function}" did not match regex: {param_str}')
    return func_data

# Builds the function declaration for a function
def get_function_declaration(func):
    decl_params = ', '.join([f'{p_type} {p_name}' for p_name, p_type in func['params'].items()])
    return f'{" ".join(func["attributes"])} {func["return_type"]} {func["name"]}({decl_params})'.strip()

# Extracts functions from source files by calling the 'get_source_context.py' script
def get_source_context(cmd, source_files):
    try:
        return subprocess.run([sys.executable, 'get_source_context.py', cmd, *source_files], cwd='..', check=True, stdout=subprocess.PIPE).stdout.decode().strip().split('\n')
    except subprocess.CalledProcessError:
        eprint('Failed to extract functions from source')
        sys.exit(1)

# Searches the source files for "#define func_name other_func_name" statements
# and returns them as a dict
RE_DEFINE_LINE = r'^\s*#define\s+(\w+)\s+(\w+)\s*$'
def find_redefinitions(source_files):
    redefinitions = {}
    for source_file in source_files:
        with open(source_file, 'r') as f:
            for line in f:
                match = re.fullmatch(RE_DEFINE_LINE, line)
                if match:
                    redefinitions[match.group(1)] = match.group(2)
    return redefinitions

# Writes functions to an excel sheet
def write_functions_to_excel(filename, funcs, prefill_params=False):
    # Create an excel document
    wb = openpyxl.Workbook()
    sheet_funcs = wb.create_sheet('functions')
    sheet_funcs.title = 'functions'
    sheet_funcs.column_dimensions['A'].width = 120
    for i in range(6):
        sheet_funcs.column_dimensions[chr(ord('B')+i)].width = 25
    sheet_snips = wb.create_sheet('snippets')
    sheet_snips.title = 'snippets'
    wb.remove(wb['Sheet'])
    # Add highlighting for void/non-void functions
    fill_green = PatternFill(start_color='99ff99', fill_type='solid')
    fill_orange = PatternFill(start_color='ffd699', fill_type='solid')
    #dxf_green = DifferentialStyle(fill=fill_green)
    #dxf_orange = DifferentialStyle(fill=fill_orange)
    #rule_void = Rule(type='containsText', operator='containsText', text='()', dxf=dxf_green)
    #rule_void.formula = ['NOT(ISERROR(SEARCH("()",A1)))']
    #rule_params = Rule(type='notContainsText', operator='notContains', text='()', dxf=dxf_orange)
    #rule_params.formula = ['ISERROR(SEARCH("()",A1))']
    #ws.conditional_formatting.add('A1:A1000', rule_void)

    for idx in range(len(funcs)):
        # Write function
        func = funcs[idx]
        sheet_funcs.cell(row=1+idx, column=1).value = get_function_declaration(func)
        sheet_funcs.cell(row=1+idx, column=1).fill = fill_green if len(func['params']) == 0 else fill_orange
        # Write function parameter names
        if prefill_params:
            p_col = 2
            for p_name in func['params']:
                sheet_funcs.cell(row=1+idx, column=p_col).value = f'{p_name}: '
                p_col += 1

    # Save document
    wb.save(filename)


def main():
    if len(sys.argv) < 2:
        eprint(f'Usage: {sys.argv[0]} <source files>... <mock header files>...')
        sys.exit(1)

    # Separate arguments into implementation source files and mock header files
    source_files = []
    header_files = []
    for file_arg in sys.argv[1:]:
        if not os.path.isfile(file_arg):
            eprint(f'{file_arg} not found')
            sys.exit(1)
        _, ext = os.path.splitext(file_arg)
        if ext.startswith('.c'):
            source_files.append(file_arg)
        elif ext.startswith('.h'):
            header_files.append(file_arg)
        else:
            eprint(f'Unknown file type: {file_arg}')
            sys.exit(1)

    # Extract all data we need from the provided files, that is:
    # - All declarations from header files that contain mock functions (and possibly other ones)
    # - All function calls inside the source files. We'll use this to filter the mock declarations.
    # - All function definitions (their declarations) in the source files to filter them out of the mock declarations
    eprint('Parsing mock declarations...')
    mock_decls_str = get_source_context('find_func_decls', header_files)
    mock_decls = [parse_function(func) for func in mock_decls_str]
    # Remove the 'extern' attribute
    for mock_decl in mock_decls:
        if 'extern' in mock_decl['attributes']:
            mock_decl['attributes'].remove('extern')
    eprint('Parsing function definitions...')
    func_calls_str = get_source_context('find_func_calls', source_files)
    func_defs_str = get_source_context('find_func_defs', source_files)
    func_defs = [parse_function(func) for func in func_defs_str]
    # Also search preprocessor definitions (we're looking for redefined function names)
    redefs = find_redefinitions(header_files)

    # Deduplicate mocks, we just take the shorter declaration for now
    mock_decls_dict = {}
    for mock_decl in mock_decls:
        if mock_decl['name'] in mock_decls_dict:
            decl_str_1 = get_function_declaration(mock_decls_dict[mock_decl['name']])
            decl_str_2 = get_function_declaration(mock_decl)
            if decl_str_1 == decl_str_2:
                pass
            elif len(decl_str_1) <= len(decl_str_2):
                eprint(f'Mock conflict: -> {decl_str_1}')
                eprint(f'Mock conflict:    {decl_str_2}')
            else:
                eprint(f'Mock conflict:    {decl_str_1}')
                eprint(f'Mock conflict: -> {decl_str_2}')
                mock_decls_dict[mock_decl['name']] = mock_decl
        else:
            mock_decls_dict[mock_decl['name']] = mock_decl
    mock_decls = list(mock_decls_dict.values())

    # Copy mock declarations when additional new names for them have been defined
    for new_name, old_name in redefs.items():
        if old_name in mock_decls_dict and new_name not in mock_decls_dict:
            mock_decl = copy.deepcopy(mock_decls_dict[old_name])
            mock_decl['name'] = new_name
            mock_decls.append(mock_decl)

    # Filter the mock declarations to keep called function only and remove functions we have implementations for
    func_defs_names = [func_def['name'] for func_def in func_defs]
    mock_decls = [mock_decl for mock_decl in mock_decls if mock_decl['name'] in func_calls_str and mock_decl['name'] not in func_defs_names]

    # Filter the function implementations to include only public functions
    # I.e. remove static functions and functions starting with <module>__ (TODO: Conti specific)
    func_defs = [func_def for func_def in func_defs if 'static' not in func_def['attributes'] and not re.match(r'[a-zA-Z0-9]+__', func_def['name'])]

    # Sort the lists first by function name, then group function with/without parameters
    mock_decls.sort(key=lambda x: x['name'])
    mock_decls.sort(key=lambda x: int(len(x['params']) == 0))
    func_defs.sort(key=lambda x: x['name'])
    func_defs.sort(key=lambda x: int(len(x['params']) == 0))

    # Write excel files
    write_functions_to_excel('testgen_mocks.xlsx', mock_decls)
    write_functions_to_excel('testgen_functions.xlsx', func_defs, prefill_params=True)


if __name__ == '__main__':
    main()
