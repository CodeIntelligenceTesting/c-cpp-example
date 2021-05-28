#!/usr/bin/env python3

import json
import openpyxl
import os
import re
import sys

def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


###########################################
# Excel file reading
###########################################

def parse_function(function):
    # Remove whitespace and trailing semicolon
    function = function.strip().strip(';')
    # For parsing make sure there is no space before the parenthesis
    function = re.sub(r' *\(', '(', function)
    # Replace (void) notation with empty parentheses
    function = re.sub(r'\( *void *\)', '()', function)
    # Parse return type, function name and parameter string
    match = re.match(r'^((?:.*[\* ])?)(\w+)\((.*)\)$', function)
    if not match:
        raise Exception(f'Function did not match regex: {function}')
    func_data = {}
    func_data['return_type'] = match.group(1).strip()
    func_data['name'] = match.group(2).strip()
    func_data['params'] = {}
    if match.group(3).strip() != '':
        params_str = [param_str.strip() for param_str in match.group(3).split(',')]
        for param_str in params_str:
            pmatch = re.match(r'^(.*[\* ])(\w+)$', param_str)
            if not pmatch:
                raise Exception(f'Parameter in function "{function}" did not match regex: {param_str}')
            func_data['params'][pmatch.group(2).strip()] = pmatch.group(1).strip()
    return func_data
        
def parse_assignment(assignment):
    param_name = assignment.split(':', 1)[0].strip()
    cmd_str = assignment.split(':', 1)[1].strip()
    match = re.match(r'^([A-Z_]+)\((.*)\)$', cmd_str)
    if not match:
        raise Exception(f'Could not parse command in assignment: {assignment}')
    cmd_args = [cmd_param.strip() for cmd_param in match.group(2).split(',')]
    return (param_name, [match.group(1)] + (cmd_args if match.group(2) != '' else []))

def read_excel_file(filepath):
    # Parse functions sheet (iterate over rows, parse column entries)
    wb = openpyxl.load_workbook(filepath)
    sheet = wb['functions']
    functions = []
    row_idx = 1
    while sheet.cell(row=row_idx, column=1).value is not None:
        # Parse function string
        func_data = parse_function(sheet.cell(row=row_idx, column=1).value)
        # Parse assignments
        func_data['assign'] = {}
        col_idx = 2
        while sheet.cell(row=row_idx, column=col_idx).value is not None:
            assignment = parse_assignment(sheet.cell(row=row_idx, column=col_idx).value)
            func_data['assign'][assignment[0]] = assignment[1]
            col_idx += 1
        functions.append(func_data)
        row_idx += 1
    # Parse snippets sheet (iterate over rows, add tuples to dict)
    snippets = {}
    if 'snippets' in wb:
        sheet_snippets = wb['snippets']
        snippet_row = 1
        while sheet_snippets.cell(row=snippet_row, column=1).value is not None:
            snippets[sheet_snippets.cell(row=snippet_row, column=1).value] = sheet_snippets.cell(row=snippet_row, column=2).value
            snippet_row += 1
    return (functions, snippets)


###########################################
# Source generation
###########################################

class FuzzDataGenerator:
    def __init__(self, prefix=''):
        self.ins_prefix = prefix
        self.gen_code = []
        # There is always at most one instance of data/size
        self.var_data = ''
        self.var_size = ''

    def gen_var_name(self, prefix):
        if not hasattr(self, 'var_counter'):
            self.var_counter = {}
        if prefix not in self.var_counter:
            self.var_counter[prefix] = 0
        var_name = f'{self.ins_prefix}{prefix}{self.var_counter[prefix]}'
        self.var_counter[prefix] += 1
        return var_name

    def gen_data(self):
        if self.var_data == '':
            self.var_data = self.gen_var_name('fuzz_data_')
            self.var_size = self.gen_var_name('fuzz_size_')
            self.gen_code.append(f'std::vector<uint8_t> {self.var_data} = fuzz_data->ConsumeBytes<uint8_t>(fuzz_data->ConsumeIntegral<uint8_t>());')
            self.gen_code.append(f'size_t {self.var_size} = {self.var_data}.size();')
        return f'{self.var_data}.data()'

    def gen_size(self):
        self.gen_data()
        return self.var_size

    def gen_size_ptr(self):
        return f'&{self.gen_size()}'

    def gen_bytes(self, size):
        var_name = self.gen_var_name('fuzz_bytes_')
        self.gen_code.append(f'uint8_t {var_name}[{size}] = {{0}};')
        self.gen_code.append(f'fuzz_data->ConsumeData({var_name}, {size});')
        return var_name

    def gen_bool(self):
        var_name = self.gen_var_name('bool_')
        self.gen_code.append(f'bool {var_name} = fuzz_data->ConsumeBool();')
        return var_name

    def gen_int(self, int_type):
        var_name = self.gen_var_name('fuzz_int_')
        self.gen_code.append(f'{int_type} {var_name} = fuzz_data->ConsumeIntegral<{int_type}>();')
        return var_name

    def gen_int_ptr(self, int_type, default=None):
        if default is None:
            return f'&' + self.gen_int(int_type)
        else:
            var_name = self.gen_var_name('fuzz_int_')
            self.gen_code.append(f'{int_type} {var_name} = {default};')
            return f'&{var_name}'

    def gen_buffer(self, size):
        var_name = self.gen_var_name('buffer_')
        self.gen_code.append(f'uint8_t {var_name}[{size}] = {{0}};')
        return var_name
    
    def gen_struct(self, name_of_struct):
        var_name = self.gen_var_name('struct_')
        self.gen_code.append(f'{name_of_struct} {var_name} = {{0}};')
        self.gen_code.append(f'fuzz_data->ConsumeData(&{var_name}, sizeof({var_name}));')
        return var_name

    def gen_struct_ptr(self, name_of_struct):
        var_name = f'&{self.gen_struct(name_of_struct)}'
        return var_name

    def gen_enum(self, name_of_enum):
        var_name = self.gen_var_name('enum_')
        self.gen_code.append(f'{name_of_enum} {var_name} = ({name_of_enum})0;')
        self.gen_code.append(f'fuzz_data->ConsumeData(&{var_name}, sizeof({var_name}));')
        return var_name

    def gen_enum_ptr(self, name_of_enum):
        var_name = f'&{self.gen_struct(name_of_enum)}'
        return var_name
    
    def gen_enum_range(self, name_of_enum, max_value):
        var_name = self.gen_var_name('enum_')
        self.gen_code.append(f'{name_of_enum} {var_name} = ({name_of_enum}) fuzz_data->ConsumeIntegralInRange(0,{max_value});')   
        return var_name
 
    def get_gen_code(self):
        return self.gen_code


def get_function_declaration(func):
    decl_params = ', '.join([f'{p_type} {p_name}' for p_name, p_type in func['params'].items()])
    return f'{func["return_type"]} {func["name"]}({decl_params})'


def generate_mocklib_function(func):
    # Iterate over assignments and generate code
    data_gen = FuzzDataGenerator()
    assign_code = []
    return_statement = ''
    for a_param, a_cmd in func['assign'].items():
        if a_param == 'return':
            if len(a_cmd) == 1 and a_cmd[0] == 'RETURN_BOOL':
                return_statement = f'return fuzz_data->ConsumeBool();'
            elif len(a_cmd) == 2 and a_cmd[0] == 'RETURN_LITERAL':
                return_statement = f'return {a_cmd[1]};'
            elif len(a_cmd) == 2 and a_cmd[0] == 'RETURN_INT':
                return_statement = f'return fuzz_data->ConsumeIntegral<{a_cmd[1]}>();'
            elif len(a_cmd) == 2 and a_cmd[0] == 'RETURN_ENUM':
                return_statement = f'return fuzz_data->ConsumeEnum<{a_cmd[1]}>();'
            elif len(a_cmd) == 2 and a_cmd[0] == 'RETURN_STRUCT':
                return_statement = f'{a_cmd[1]} return_struct;\n\rConsumeDataAndFillRestWithZeros(&return_struct, sizeof(return_struct), fuzz_data);\n\rreturn return_struct;'
            elif len(a_cmd) == 3 and a_cmd[0] == 'RETURN_ENUM_RANGE':
                return_statement = f'return static_cast<{a_cmd[1]}>(fuzz_data->ConsumeIntegralInRange<uint32_t>(0, static_cast<uint32_t>({a_cmd[2]})));'
            else:
                raise Exception(f'Invalid return command: {" ".join(a_cmd)}')
        elif a_param in func['params']:
            if len(a_cmd) == 2 and a_cmd[0] == 'WRITE_BYTES':
                if a_cmd[1].isnumeric(): # WRITE_BYTES(123)
                    bytes_size = a_cmd[1]
                elif a_cmd[1] in func['params']: # WRITE_BYTES(someVariable)
                    # If someVariable is a pointer, dereference it
                    bytes_size = f'*{a_cmd[1]}' if func['params'][a_cmd[1]].endswith('*') else a_cmd[1]
                elif "sizeof(" in a_cmd[1]: # WRITE_BYTES(sizeof(someStruct))
                    bytes_size = a_cmd[1];
                else:
                    raise Exception(f'{func["name"]}: Invalid parameter referenced in command: {" ".join(a_cmd)}')
                assign_code.append(f'fuzz_data->ConsumeData({a_param}, {bytes_size});')
            else:
                raise Exception(f'Invalid mock command: {" ".join(a_cmd)}')
        else:
            raise Exception(f'Invalid parameter "{a_param}" for function "{func["name"]}"')
    gen_code = data_gen.get_gen_code()
    # Assemble the code snippets to create a mock
    mock_code = []
    if len(gen_code) > 0:
        #mock_code.append('// Prepare fuzz data to be returned')
        mock_code.extend(gen_code)
    if len(assign_code) > 0:
        #mock_code.append('// Write and return fuzz data')
        mock_code.extend(assign_code)
    if return_statement != '':
        mock_code.append(return_statement)
    if len(mock_code) == 0:
        mock_code.append('// Nothing to do here')
    # Assemble the function definition
    signature = get_function_declaration(func)
    mock_body = '\n'.join([f'    {line}' for line in mock_code])
    return f'{signature} {{\n{mock_body}\n}}'

def generate_mocklib(data, snippets, out_dir):
    mocklib_template = """
#include "mocklib.h"
#include <vector>
#include <cstdint>
#include <iostream>
#include <iomanip>

extern "C" {
%INCLUDES%
}

%GLOBAL_VARS%

static FuzzedDataProvider *fuzz_data;

// Wrapper function for FuzzedDataProvider.h
// Writes |num_bytes| of input data to the given destination pointer. If there
// is not enough data left, writes all remaining bytes and fills the rest with zeros.
// Return value is the number of bytes written.
void ConsumeDataAndFillRestWithZeros(void *destination,
                                              size_t num_bytes, FuzzedDataProvider *fuzz_data) {
  size_t num_bytes_with_fuzz_data = fuzz_data->ConsumeData(destination, num_bytes);
  if (num_bytes > num_bytes_with_fuzz_data) {
    size_t num_bytes_with_zeros = num_bytes - num_bytes_with_fuzz_data;
    std::memset((char*)destination+num_bytes_with_fuzz_data, 0, num_bytes_with_zeros);
  }
}


void mocklib_set_data(void *fuzzed_data_provider) {
    fuzz_data = (FuzzedDataProvider *) fuzzed_data_provider;
}

%DEFINITIONS%
"""
    mocklib_header_template = """
#ifndef MOCKLIB_H
#define MOCKLIB_H

#include "FuzzedDataProvider.h"
#ifdef __cplusplus
extern "C" {
#endif

%INCLUDES%

void ConsumeDataAndFillRestWithZeros(void *destination, size_t num_bytes, FuzzedDataProvider *fuzz_data);

void mocklib_set_data(void *fuzzed_data_provider);

%DECLARATIONS%

#ifdef __cplusplus
}
#endif

#endif // MOCKLIB_H
"""
    # Iterate over functions, generate function bodies and collect declarations
    declarations = []
    mocks = []
    for func in data:
        mock_str = generate_mocklib_function(func)
        mocks.append(mock_str)
        declarations.append(get_function_declaration(func) + ';')
    # Create mocklib.cpp file
    mocks_str = '\n\n'.join(mocks)
    includes_str = snippets['includes'] if 'includes' in snippets else ''
    globals_str = snippets['global_vars'] if 'global_vars' in snippets else ''
    mocklib_cpp = mocklib_template.replace('%DEFINITIONS%', mocks_str)
    mocklib_cpp = mocklib_cpp.replace('%INCLUDES%', includes_str)
    mocklib_cpp = mocklib_cpp.replace('%GLOBAL_VARS%', globals_str)
    with open(os.path.join(out_dir, 'mocklib.cpp'), 'w') as f:
        f.write(mocklib_cpp)
    # Create mocklib.h file
    decls_str = '\n'.join(declarations)
    mocklib_h = mocklib_header_template.replace('%DECLARATIONS%', decls_str)
    mocklib_h = mocklib_h.replace('%INCLUDES%', includes_str)
    with open(os.path.join(out_dir, 'mocklib.h'), 'w') as f:
        f.write(mocklib_h)


def generate_fuzztest_invocation(func):
    call_param_values = []
    data_gen = FuzzDataGenerator()
    for p_name, p_type in func['params'].items():
        if p_name not in func['assign']:
            raise Exception(f'{func["name"]}: Parameter "{p_name}" has no value assigned')
        cmd = func['assign'][p_name]
        if len(cmd) == 2 and cmd[0] == 'ARG_LITERAL':
            call_param_values.append(cmd[1])
        elif len(cmd) == 1 and cmd[0] == 'ARG_DATA':
            call_param_values.append(data_gen.gen_data())
        elif len(cmd) == 1 and cmd[0] == 'ARG_SIZE':
            call_param_values.append(data_gen.gen_size())
        elif len(cmd) == 1 and cmd[0] == 'ARG_SIZE_PTR':
            call_param_values.append(data_gen.gen_size_ptr())
        elif len(cmd) == 2 and cmd[0] == 'ARG_BYTES':
            call_param_values.append(data_gen.gen_bytes(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_INT':
            call_param_values.append(data_gen.gen_int(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_INT_PTR':
            call_param_values.append(data_gen.gen_int_ptr(cmd[1]))
        elif len(cmd) == 3 and cmd[0] == 'ARG_INT_PTR':
            call_param_values.append(data_gen.gen_int_ptr(cmd[1], default=cmd[2]))
        elif len(cmd) == 1 and cmd[0] == 'ARG_BOOL':
            call_param_values.append(data_gen.gen_bool())
        elif len(cmd) == 2 and cmd[0] == 'ARG_BUFFER':
            call_param_values.append(data_gen.gen_buffer(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_STRUCT':
            call_param_values.append(data_gen.gen_struct(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_STRUCT_PTR':
            call_param_values.append(data_gen.gen_struct_ptr(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_ENUM':
            call_param_values.append(data_gen.gen_enum(cmd[1]))
        elif len(cmd) == 2 and cmd[0] == 'ARG_ENUM_PTR':
            call_param_values.append(data_gen.gen_enum_ptr(cmd[1]))
        elif len(cmd) == 3 and cmd[0] == 'ARG_ENUM_RANGE':
            call_param_values.append(data_gen.gen_enum_range(cmd[1], cmd[2]))
        else:
            raise Exception(f'{func["name"]}: Parameter "{p_name}" has unknown command: {cmd}')
    func_invocation = f'{func["name"]}({", ".join(call_param_values)});'
    global print_info
    if print_info:
        print_call_info = f'printf("Calling Function {func["name"]}\\n")'
        return data_gen.get_gen_code() + [print_call_info] + [func_invocation]
    else:
        return data_gen.get_gen_code() + [func_invocation]

def generate_fuzztests(data, snippets, out_dir):
    fuzztest_template = """
#include <stdint.h>
#include <stddef.h>

#include "FuzzedDataProvider.h"
#include "mocklib.h"

// Includes and function declarations
extern "C" {
%INCLUDES%

%DECLARATIONS%
}

extern "C" int FUZZ(const uint8_t *Data, size_t Size) {
    // Setup FuzzedDataProvider and initialize the mocklib
    FuzzedDataProvider fuzz_data_provider(Data, Size);
    FuzzedDataProvider *fuzz_data = &fuzz_data_provider;
    mocklib_set_data(fuzz_data);

    int number_of_functions = fuzz_data->ConsumeIntegralInRange<int>(1,100);
    for (int i=0; i<number_of_functions; i++) {
%FUZZTEST%
    }

    return 0;
}
"""
    # Generate function declarations
    func_decls_source = '\n'.join([get_function_declaration(func) + ';' for func in data])
    # Generate code for function invocations in a switch case statement
    invocations = []
    for func in data:
        if len(func['params']) == 0:
            invocation_code = []
            global print_info
            if print_info:
                print_call_info = f'printf("Calling Function {func["name"]}\\n")'
                invocation_code = [print_call_info]
            invocation_code += [f'{func["name"]}();']
        else:
            invocation_code = generate_fuzztest_invocation(func)
        invocations.append(invocation_code)
    switch_source = f'      int func_id = fuzz_data->ConsumeIntegralInRange<int>(0, {len(invocations) - 1});\n'
    switch_source += '      switch(func_id) {\n'
    for func_id in range(len(invocations)):
        switch_source += f'         case {func_id}: {{\n'
        switch_source += '\n'.join([(' ' * 12) + line for line in invocations[func_id]]) + '\n'
        switch_source += (' ' * 12) + 'break;\n'
        switch_source += '          }\n'
    switch_source += '      }'
    # Generate fuzz test source from template
    includes_str = snippets['includes'] if 'includes' in snippets else ''
    fuzztest_source = fuzztest_template.replace('%INCLUDES%', includes_str)
    fuzztest_source = fuzztest_source.replace('%DECLARATIONS%', func_decls_source)
    fuzztest_source = fuzztest_source.replace('%FUZZTEST%', switch_source)
    with open(os.path.join(out_dir, 'fuzztest.c'), 'w') as f:
        f.write(fuzztest_source)


###########################################
# Main
###########################################

def main():
    # Check usage and get arguments
    if len(sys.argv) < 4:
        eprint('Usage: {} mocklib|fuzztests <excel sheet> <output directory> (print call info (1|0))')
        sys.exit(1)

    arg_cmd = sys.argv[1]
    arg_excel_file = sys.argv[2]
    arg_out_dir = sys.argv[3]

    global print_info 
    print_info = False
 
    if len(sys.argv) > 4:
        if sys.argv[4] == '1':
            print_info = True
        else:
            print_info = False

    functions, snippets = read_excel_file(arg_excel_file)
    #print(json.dumps(excel_data[0], indent=4))

    if arg_cmd == 'mocklib':
        generate_mocklib(functions, snippets, arg_out_dir)
    elif arg_cmd == 'fuzztests':
        generate_fuzztests(functions, snippets, arg_out_dir)


if __name__ == '__main__':
    main()
